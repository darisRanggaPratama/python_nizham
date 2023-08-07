import openpyxl

def check_duplicate_cells_in_workbook(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Create a dictionary to store cell values, sheet names, and cell addresses
    cell_values_info = {}

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Iterate through each cell in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                cell_value = cell.value
                if cell_value is not None:
                    cell_address = cell.coordinate
                    cell_info = (sheet_name, cell_address)

                    if cell_value in cell_values_info:
                        cell_values_info[cell_value].append(cell_info)
                    else:
                        cell_values_info[cell_value] = [cell_info]

    # Filter the dictionary to keep only the duplicate cell values and their info
    duplicate_cell_values_info = {value: info_list for value, info_list in cell_values_info.items() if len(info_list) > 1}

    return duplicate_cell_values_info

if __name__ == "__main__":
    file_path = "blank_cell.xlsx"  # Update with the correct file path

    duplicate_cells_info = check_duplicate_cells_in_workbook(file_path)
    if duplicate_cells_info:
        print("Duplicate cell values found: ;")
        for value, info_list in duplicate_cells_info.items():
            for sheet_name, cell_address in info_list:
                print(f"Sheet: {sheet_name} ; Value: {value} ; Cell: {cell_address}")
    else:
        print("No duplicate cell values found.")
