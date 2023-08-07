from openpyxl import load_workbook

path = 'blank_cell.xlsx'

wb = load_workbook(path)

sheet = wb['empty']

col = int(input("Enter column number: "))

for rowNum in range(1, sheet.max_row):
    if((sheet.cell(row=rowNum, column=col).value)==None):
        print(f"The first empty cell from column {col} is:",sheet.cell(row=rowNum, column=col))
        break

# This code is work



    