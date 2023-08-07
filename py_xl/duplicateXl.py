import openpyxl

def check_duplicate_cells_in_workbook(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Create a dictionary to store cell values as keys and their addresses as values
    cell_values_and_addresses = {}

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Iterate through each cell in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                cell_value = cell.value
                if cell_value is not None:
                    if cell_value in cell_values_and_addresses:
                        cell_values_and_addresses[cell_value].append(cell.coordinate)
                    else:
                        cell_values_and_addresses[cell_value] = [cell.coordinate]

    # Filter the dictionary to keep only the duplicate cell values and their addresses
    duplicate_cell_values_and_addresses = {value: addresses for value, addresses in cell_values_and_addresses.items() if len(addresses) > 1}

    return duplicate_cell_values_and_addresses

if __name__ == "__main__":
    file_path = "btu_bms.xlsx"  # Update with the correct file path

    duplicate_cells = check_duplicate_cells_in_workbook(file_path)
    if duplicate_cells:
        print("\nDuplicate cell values found:")
        for value, addresses in duplicate_cells.items():
            print(f"Value: {value} ; Cell: {', '.join(addresses)}")
    else:
        print("No duplicate cell values found.")
        

# This code is works
