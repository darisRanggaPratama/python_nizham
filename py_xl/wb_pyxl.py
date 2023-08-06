from openpyxl import Workbook

wb = Workbook()

ws = wb.active

ws1 = wb.create_sheet("Pemasukan")
ws2 = wb.create_sheet("Pemasukan", 0)
ws3 = wb.create_sheet("Pengeluaran", -1)

print(wb.sheetnames)

wb.save('wb_pyxl.xls')
