from openpyxl import *
from openpyxl.utils import *



get_column_letter(1)
print('huruf-kolom ke 1: ', get_column_letter(1))
print('huruf-kolom ke 5: ', get_column_letter(5))

wb = load_workbook('blank_cell.xlsx')
sh_data = wb['DATA']
get_column_letter(sh_data.max_column)
print('max kolom: ', get_column_letter(sh_data.max_column))

# print(f'huruf-kolom ke angka: {column_index_from_string('B')} {column_index_from_string('G')}')
print('kolom B: ', column_index_from_string('B'))
print('kolom G: ', column_index_from_string('G'))