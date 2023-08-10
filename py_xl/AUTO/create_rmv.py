from openpyxl import *

wbook = Workbook()
wbook.sheetnames
print('1', wbook.sheetnames)
wbook.create_sheet()
print('2', wbook.create_sheet())
wbook.sheetnames
print('3', wbook.sheetnames)
# Create sheet at index 0
wbook.create_sheet(index=0, title='First')
print('4', wbook.sheetnames)
wbook.create_sheet(index=1, title='second')
print('5', wbook.sheetnames)
# Delete sheet
del wbook['First']
print('6', wbook.sheetnames)

