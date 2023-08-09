import openpyxl

path = 'blank_cell.xlsx'
wb = openpyxl.load_workbook(path)
print(type(wb))

wb.sheetnames
# Menampilkan semua worksheet/ sheet dalam workbook
print(wb.sheetnames) 

# Mengambil 1 sheet
sheet = wb['GAJI']
print(sheet)
print(type(sheet))

# Nama Sheet
sheet.title
print(sheet.title)

# Memilih sheet active
activeSheet = wb.active
print(activeSheet)



# Fungsi load_workbook dari modul openpyxl digunakan untuk membuka file Excel. Setelah file dimuat, Anda dapat berinteraksi dengan data di dalamnya, seperti membaca atau memodifikasi sel.