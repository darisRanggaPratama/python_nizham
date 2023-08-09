from openpyxl import *

path = 'blank_cell.xlsx'
wb = load_workbook(path)
sh_data = wb['DATA']

# Cell tertentu
sh_data.cell(row=1, column=2)
print('cell: ', sh_data.cell(row=1, column=2))

sh_data.cell(row=1, column=2).value
print('nilai cell: ', sh_data.cell(row=1, column=2).value)

for x in range(1, 10, 2):
    print(x, sh_data.cell(row=x, column=2).value)
    
    
sh_data.max_row
print('Jumlah Baris: ', sh_data.max_row)

sh_data.max_column
print('Jumlah Kolom: ', sh_data.max_column)
    
# for x in range(1, 10, 2):
# Ini adalah loop for yang akan mengiterasi melalui angka-angka dari 1 hingga 9 dengan loncatan 2 setiap kali iterasi. Jadi, nilai x akan menjadi 1, 3, 5, 7, dan 9.

# print(x, sh_data.cell(row=x, column=2).value)
# Di setiap iterasi loop, ini akan mencetak nilai x (yang merupakan baris ke-x) dan nilai yang terdapat di kolom ke-2 (kolom B) dari sheet sh_data.

# sh_data.cell(row=x, column=2) mengambil objek sel (cell) dari baris x dan kolom kedua.
# .value digunakan untuk mengambil nilai yang terdapat di dalam sel tersebut.