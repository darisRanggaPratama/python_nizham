from openpyxl import *

path = 'blank_cell.xlsx'
wb = load_workbook(path)
print(wb.sheetnames)
sheet = wb['DATA']

# Ambil cell
sheet['A1']
# print(sheet['A1'])

# Ambil nilai di cell
sheet['A1'].value
print(sheet['A1'].value)

val = sheet['B1']
val.value
print(val.value)

# Ambil baris, kolom & nilai dari cell
print(f'huruf-kolom: {utils.get_column_letter(val.column)} baris: {val.row} kolom: {val.column} \nCell: {val.coordinate} nilai:  {val.value}')

# Ambil nilai
print('Isi B1: ', sheet['B1'].value)