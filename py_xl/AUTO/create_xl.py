from openpyxl import *

# Create Workbook
wbook = Workbook()
wbook.sheetnames
# print(wbook.sheetnames)
wsheet = wbook.active
wsheet.title
print(wsheet.title)
wsheet.title = 'kuliner'
print(wbook.sheetnames)
