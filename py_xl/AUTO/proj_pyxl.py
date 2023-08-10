import openpyxl
import pprint

print('Opening Workbook...')
path = 'censuspopdata.xlsx'
wbook = openpyxl.load_workbook(path)

wsheet = wbook['Population by Census Tract']
countyData = {}

print('Reading rows..')
for row in range(2, wsheet.max_row + 1):
    state = wsheet['B' + str(row)].value
    county = wsheet['C' + str(row)].value
    pop = wsheet['D' + str(row)].value
    
    countyData.setdefault(state, {})
    countyData[state].setdefault(county, {'tracts':0, 'pop': 0})
    countyData[state][county]['tracts'] += 1
    countyData[state][county]['pop'] += int(pop)
    
print('Writing results...')
resultFile = open('census2010.py', 'w')
resultFile.write('allData = ' + pprint.pformat(countyData))
resultFile.close()
print('Done...')
    

# Inisialisasi Dictionary untuk Menyimpan Data:

# countyData = {}

# Variabel countyData adalah dictionary yang akan digunakan untuk menyimpan data tentang populasi dan traktat sensus berdasarkan negara bagian dan kabupaten.

# Loop Melalui Baris Data:

# for row in range(2, wsheet.max_row + 1):
#     state = wsheet['B' + str(row)].value
#     county = wsheet['C' + str(row)].value
#     pop = wsheet['D' + str(row)].value

# Loop ini akan berjalan melalui setiap baris data mulai dari baris kedua hingga baris terakhir di lembar kerja. Variabel state akan berisi nilai dari kolom B (negara bagian), variabel county akan berisi nilai dari kolom C (kabupaten), dan variabel pop akan berisi nilai dari kolom D (populasi).

# Mengisi Data ke Dalam Dictionary:

# countyData.setdefault(state, {})
# countyData[state].setdefault(county, {'tracts':0, 'pop': 0})
# countyData[state][county]['tracts'] += 1
# countyData[state][county]['pop'] += int(pop)

# Pertama, kode di atas memastikan bahwa ada entri untuk negara bagian dan kabupaten tertentu dalam dictionary countyData.
# Selanjutnya, kode mengatur entri default dengan jumlah traktat dan populasi awal (0) jika belum ada di dictionary.
# Kode kemudian menambahkan 1 pada jumlah traktat dan menambahkan nilai populasi ke entri yang sesuai.

# Menulis Hasil ke File Python:

# resultFile = open('census2010.py', 'w')
# resultFile.write('allData = ' + pprint.pformat(countyData))
# resultFile.close()

# Kode di atas membuka file dengan nama 'census2010.py' untuk ditulis ('w'). Kemudian, kode menulis string yang berisi representasi formatted dari countyData menggunakan pustaka pprint. Akhirnya, file ditutup.

# Hasil akhirnya adalah file 'census2010.py' yang berisi representasi dictionary countyData dalam format yang mudah dibaca. Anda dapat mengimpor file ini dalam program Python lain untuk mengakses data yang telah diproses.
