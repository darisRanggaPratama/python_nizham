import openpyxl

def check_duplicate_cells_in_workbook(file_path):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)

    # Create a set to store unique cell values and a list to store duplicate cell values
    unique_cell_values = set()
    duplicate_cell_values = []

    # Iterate through each sheet in the workbook
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        
        # Iterate through each row in the sheet
        for row in sheet.iter_rows(values_only=True):
            # Iterate through each cell in the row
            for cell_value in row:
                if cell_value in unique_cell_values:
                    duplicate_cell_values.append((sheet_name, cell_value))
                else:
                    unique_cell_values.add(cell_value)

    # Return the list of duplicate cell values
    return duplicate_cell_values

if __name__ == "__main__":
    file_path = "blank_cell.xlsx"  # Update with the correct file path

    duplicates = check_duplicate_cells_in_workbook(file_path)
    if duplicates:
        print("Duplicate cell values found:")
        for sheet_name, cell_value in duplicates:
            print(f"Sheet: {sheet_name}, Value: {cell_value}")
    else:
        print("No duplicate cell values found.")

# This code is works