from openpyxl import Workbook

wb = Workbook()
ws = wb.active
sem = ['kami', 'sedang', 'belajar', 'python']

for row in range(0, len(sem)):
	ws.cell(row = row + 1, column=1).value= sem[row]


for column in range(0, len(sem)):
	ws.cell(row = 1, column = column+1).value = sem[column]


wb.save('cell_pyxl.xls')