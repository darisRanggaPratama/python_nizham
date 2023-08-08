import openpyxl

def get_cell_coordinates(file_path, sheet_name, target_value):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[sheet_name]
    
    # Search for the target value in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        for column_index, cell_value in enumerate(row, start=1):
            if cell_value == target_value:
                column_letter = openpyxl.utils.get_column_letter(column_index)
                row_number = row[0]
                cell_coordinates = f"{column_letter} {row_number}"
                return cell_coordinates
    
    return None  # Target value not found

file_path = "blank_cell.xlsx"
sheet_name = "DATA"  # Change this to your desired sheet name
target_value = "MINEVA"

cell_coordinates = get_cell_coordinates(file_path, sheet_name, target_value)

if cell_coordinates:
    print(f"The cell containing '{target_value}' is located at: {cell_coordinates}")
else:
    print(f"'{target_value}' not found in the worksheet.")
