import openpyxl
from collections import defaultdict
from openpyxl.utils import column_index_from_string


def find_duplicate_cells(file_path, sheet_name, column_letter):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[sheet_name]
    
    # Convert column letter to index
    column_index = column_index_from_string(column_letter)
    
    # Create a dictionary to store cell values and their addresses
    cell_dict = defaultdict(list)
    
    # Iterate through the rows in the specified column
    for row in worksheet.iter_rows(min_row=2, min_col=column_index, values_only=True):
        for cell in row:            
            cell_value = row[0]
            cell_address = f"{column_letter} {row[0]}"
            cell_dict[cell_value].append(cell_address)
    
    # Find duplicate cell values
    duplicate_cells = {value: addresses for value, addresses in cell_dict.items() if len(addresses) > 1}
    
    return duplicate_cells

def get_cell_coordinates(file_path, sheet_name, target_value):
    # Load the Excel workbook
    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook[sheet_name]
    
    # Search for the target value in the worksheet
    for row in worksheet.iter_rows(values_only=True):
        for column_index, cell_value in enumerate(row, start=1):
            if cell_value == target_value:
                column_letter = openpyxl.utils.get_column_letter(column_index)
                row_number = row[0]
                cell_coordinates = f"{column_letter}{row_number}"
                return cell_coordinates
    
    return None  # Target value not found

file_path = "blank_cell.xlsx" # "path/to/your/excel/file.xlsx"
sheet_name = "DATA"  # Change this to your desired sheet name
column_letter = "F"    # Change this to your desired column letter

duplicate_cells = find_duplicate_cells(file_path, sheet_name, column_letter)

# Print the duplicate cell values along with their addresses
for value, addresses in duplicate_cells.items():
    print(f"Duplicate value: {value}")
    for address in addresses:
        print(f"Sheet: {sheet_name}, Column: {address}")
    print("="*30)

    target_value = value

    cell_coordinates = get_cell_coordinates(file_path, sheet_name, target_value)

    if cell_coordinates:
        print(f"The cell containing '{target_value}' is located at: {cell_coordinates}")
    else:
        print(f"'{target_value}' not found in the worksheet.")








