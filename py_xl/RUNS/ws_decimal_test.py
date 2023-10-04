from openpyxl import *
from openpyxl.utils import *

# Isi Nama File
#path = input('\nDecimal Value in Worksheet\nFile Name: ')

# Akses file Excel
wbook = load_workbook('sample.xlsx')

# Tampilkan Worksheet tersedia
# print(f'Available Worksheet:\n{wbook.sheetnames}')

# Pilih worksheet yang akan Anda periksa
# wsheet = input('Get 1 Worksheet: ')
sheet = wbook['GAJI']

# Maksimal kolom terisi nilai
max = get_column_letter(sheet.max_column)

print(f'\nSearch in: ;\nDECIMAL VALUE ;\n\nNo ; Cell ; Value ;')

# Huruf-Kolom ke Angka-Kolom
start = 0
end = column_index_from_string(max)
x = 0
for start in range(end):
    # Cek decimal value di kolom 1 per 1
    range_col = list(sheet.columns)[start]
    for cell in range_col:
        if isinstance(cell.value, float):
            x += 1
            print(f' {x} ; {cell.coordinate} ; {cell.value} ;')

print('\n===End Searching===;')
wbook.close()
