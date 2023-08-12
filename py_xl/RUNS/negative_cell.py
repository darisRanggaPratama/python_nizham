from openpyxl import *
from openpyxl.utils import *

# Isi Nama File
path = input('\nNEGATIVE CELLS\nFile Name: ')

# Akses file excel
wbook = load_workbook(path)

# Tampilkan worksheet tersedia
print(f'Available Worksheet: {wbook.sheetnames}')
# Pilih Worksheet
wsheet = input('Get 1 Worksheet: ')

# Akses worksheet tertentu
sheet = wbook[wsheet]
# Kolom yang tersedia dalam worksheet
max = get_column_letter(sheet.max_column)
print(f'Maximal Column: {max}')
colm = input('Get 1 Column: ')

# Akses kolom tertentu
col = column_index_from_string(colm)
range_col = list(sheet.columns)[col-1]

print('\nNEGATIVE CELLS ;\n\nNo ; Cell ; Value ; IsNegative')
# Baca nilai di kolom
x = 0
for cell in range_col:
    # Cek blank value 1 per 1
    if isinstance(cell.value, (int, float)):
        negative = cell.value < 0
        if negative is True:
            x += 1
            # Bila cell kosong tampilkan
            print(f'{x} ; {cell.coordinate} ; {cell.value} ; {negative} ;')


print('\n===End Searching===;')
wbook.close()
