import openpyxl

# Buka file Excel
file_path = 'blank_cell.xlsx'  # Ganti dengan path file Excel Anda
workbook = openpyxl.load_workbook(file_path)

# Loop melalui semua sheet di workbook
print("FLOATING VALUE ; \nSHEET ; CELL ; VALUE ;")
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    
    # Loop melalui semua sel di sheet
    for row in sheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, float):
                float_value = cell.value
                cell_address = cell.coordinate
                print(f"{sheet_name} ; {cell_address} ; {float_value} ;")

# Tutup file Excel setelah selesai
workbook.close()

# THIS CODE IS WORK