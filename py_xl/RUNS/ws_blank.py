from openpyxl import *
from openpyxl.utils import *

# Isi Nama File
path = input('\nBlank/ Empty/ Null Value in Worksheet\nFile Name: ')

# Akses file Excel
wbook = load_workbook(path)

# Tampilkan Worksheet tersedia
print(f'Available Worksheet:\n{wbook.sheetnames}')

# Pilih worksheet yang akan Anda periksa
wsheet = input('Get 1 Worksheet: ')
sheet = wbook[wsheet]

# Maksimal kolom terisi nilai
max = get_column_letter(sheet.max_column)

print(f'\nSearch in: {wsheet};\nNULL VALUE ;\n\nNo ; Cell ; Value ;')

# Huruf-Kolom ke Angka-Kolom
start = 0
end = column_index_from_string(max)
x = 0
for start in range(end):
    # Cek nilai: NULL di kolom 1 per 1
    range_col = list(sheet.columns)[start]
    for cell in range_col:
        # Cek cell value = NULL
        if cell.value == None :
            x += 1
            # Bila ada: NULL tampilkan
            print(f' {x} ; {cell.coordinate} ; {cell.value} ;')


print('\n===End Searching===;')
wbook.close()
