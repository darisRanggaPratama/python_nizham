from openpyxl import *
from openpyxl.utils import *

# Isi Nama File
path = input('\nZERO VALUE IN CELLS\nFile Name: ')

# Akses file excel
wbook = load_workbook(path)

# Tampilkan worksheet tersedia
print(f'Available Worksheet: {wbook.sheetnames}')
# Pilih Worksheet
wsheet = input('Get 1 Worksheet: ')

# Akses worksheet tertentu
sheet = wbook[wsheet]
# Kolom yang tersedia dalam worksheet
max = get_column_letter(sheet.max_column)
print(f'Maximal Column: {max}')
colm = input('Get 1 Column: ')

# Akses kolom tertentu
col = column_index_from_string(colm)
range_col = list(sheet.columns)[col-1]

print('\nZERO VALUE IN CELLS ;\n\nNo ; Cell ; Value ;')
# Baca nilai di kolom
x = 0
for cell in range_col:
    # Cek zero value di cell 1 per 1
    if isinstance(cell.value, (int, float)):
        # Cek cell value: 0
        if -1 < cell.value < 1:
            x += 1
            # Bila cell negative tampilkan
            print(f'{x} ; {cell.coordinate} ; {cell.value} ;')


print('\n===End Searching===;')
wbook.close()
