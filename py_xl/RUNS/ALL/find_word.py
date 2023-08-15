from openpyxl import *

# Nama File
path = input('\nFIND WORD\nFile Name: ')

# Kata yang dicari
words = input('Word: ')
words_lower = words.lower()

# Akses workbook Excel
wbook = load_workbook(path)
# Tampilkan worksheet tersedia
wsheet = wbook.sheetnames
print(f'\nAvailable Worksheet: \n{wsheet}')


# Fungsi untuk mencari kata dalam workbook
def find_word(worksheet, word):
    for row in worksheet.iter_rows():
        for cell in row:
            # Cek isi cell & kata dicari
            if isinstance(cell.value, str) and word in cell.value.lower():
                return cell


# Cari nilai dalam setiap worksheet
print('\nNo ; Sheet ; Cell ; Value ;')
print('-' * 40)
x = 0
for sheets in wsheet:
    # Akses worksheet tertentu
    sheet = wbook[sheets]
    cells = find_word(sheet, words_lower)
    if cells:
        x += 1
        print(f'{x} ; {sheets} ; {cells.coordinate} ; {cells.value} ;')

# Tutup workbook
wbook.close()
