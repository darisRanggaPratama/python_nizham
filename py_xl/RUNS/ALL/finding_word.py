import openpyxl

# Buka workbook Excel
workbook = openpyxl.load_workbook('sample.xlsx')  # Ganti 'nama_file.xlsx' dengan nama file Anda
sheet_names = workbook.sheetnames

# Kata yang ingin dicari dalam sel
kata_tertentu = 'MIKASA'  # Ganti 'contoh_kata' dengan kata yang ingin Anda cari
kata_tertentu_lower = kata_tertentu.lower()  # Konversi kata menjadi lowercase

# Fungsi untuk mencari nilai dalam worksheet (case insensitive)
def cari_nilai(worksheet, kata):
    for row in worksheet.iter_rows():
        for cell in row:
            if isinstance(cell.value, str) and kata in cell.value.lower():
                return cell

# Cari nilai dalam setiap worksheet
for sheet_name in sheet_names:
    sheet = workbook[sheet_name]
    cell_with_value = cari_nilai(sheet, kata_tertentu_lower)
    if cell_with_value:
        print("Worksheet:", sheet_name)
        print("Alamat Cell:", cell_with_value.coordinate)
        print("Nilai Cell:", cell_with_value.value)
        print('-' * 40)

# Tutup workbook
workbook.close()
