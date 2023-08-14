from openpyxl import *
from openpyxl.utils import *

def duplicate(file_path):
    # Akses file excel
    wbook = load_workbook(file_path)
    # Tampilkan Worksheet tersedia
    print(f'Available Worksheet:\n{wbook.sheetnames}')

    # Pilih Worksheet
    wsheet = input('Get 1 Worksheet: ')
    # Akses Worksheet
    sheet = wbook[wsheet]

    # Maksimal kolom
    max = get_column_letter(sheet.max_column)
    print(f'Maximal Column: {max}')
    # Pilih Kolom
    colm = input('Get 1 Column: ')

    # Akses kolom tertentu
    col = column_index_from_string(colm)
    range_col = list(sheet.columns)[col-1]

    # Dictionary
    cell_values_info = {}

    # Iterasi nilai cel dari kolom
    for cell in range_col:
        cell_value = cell.value
        # Jika nilai cell tidak kosong
        if cell_value is not None:
            cell_address = cell.coordinate
            cell_info = cell_address

            if cell_value in cell_values_info:
                cell_values_info[cell_value].append(cell_info)
            else:
                cell_values_info[cell_value] = [cell_info]

    duplicate_cell_values_info = {
        value: info_list for value,
        info_list in cell_values_info.items() if len(info_list) > 1
    }

    wbook.close()
    return duplicate_cell_values_info


# Isi Nama File
file_path = input('\nDUPLICATE VALUE\nFile Name: ')

# Jalankan fungsi lalu return ditampung di variable
duplicate_cell_info = duplicate(file_path)

if duplicate_cell_info:
    print('\nDUPLICATE CELL ;\n\nNo ; Cell ; Values ;')
    x = 0
    # Tampilkan nilai duplikat
    for value, info_list in duplicate_cell_info.items():
        for cell_address in info_list:
            x += 1
            print(f'{x} ; {cell_address} ; {value} ;')
    print('===End Searching===;')





