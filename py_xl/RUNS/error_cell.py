from openpyxl import *
from openpyxl.utils import *

# Isi Nama File
path = input('\nERROR Value\nFile Name: ')

# Akses file excel
wbook = load_workbook(path)

# Tampilkan Worksheet tersedia
print(f'Available Worksheet: {wbook.sheetnames}')
# Pilih Worksheet
wsheet = input('Get 1 Worksheet: ')

# Akses worksheet tertentu
sheet = wbook[wsheet]
# Kolom yang tersedia dalam wosksheet
max = get_column_letter(sheet.max_column)
print(f'Maximal Column: {max}')
# Pilih kolom
colm = input('Get 1 Column: ')

# Akses kolom tertentu
col = column_index_from_string(colm)
range_col = list(sheet.columns)[col-1]

print('\nERROR VALUE ;\n\nNo ; Cell ; Value ;')
# Baca nilai di kolom
x = 0
for cell in range_col:
    # Cek karakter: # value 1 per 1
    if isinstance(cell.value, str):
        # Cek panjang cell value
        if len(cell.value) > 0:
            # Cek karakter dicari
            if cell.value[0] == '#':
                x += 1
                # Bila ada: # tampilkan
                print(f' {x} ; {cell.coordinate} ; {cell.value} ;')


print('\n===End Searching===;')
wbook.close()
