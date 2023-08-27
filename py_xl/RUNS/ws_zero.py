from openpyxl import *
from openpyxl.utils import *

# Isi Nama File
path = input('\nZero Value in Worksheet\nFile Name: ')

# Akses file Excel
wbook = load_workbook(path)

# Tampilkan Worksheet tersedia
print(f'Available Worksheet:\n{wbook.sheetnames}')

# Pilih worksheet yang akan Anda periksa
wsheet = input('Get 1 Worksheet: ')
sheet = wbook[wsheet]

# Maksimal kolom terisi nilai
max = get_column_letter(sheet.max_column)

print(f'\nSearch in: {wsheet};\nZERO VALUE ;\n\nNo ; Cell ; Value ;')

# Huruf-Kolom ke Angka-Kolom
start = 0
end = column_index_from_string(max)
x = 0
for start in range(end):
    # Cek nilai: 0 di kolom 1 per 1
    range_col = list(sheet.columns)[start]
    for cell in range_col:
        if isinstance(cell.value, (int, float)):
            # Cek panjang cell value
            if -1 < cell.value < 1:
                x += 1
                # Bila ada: 0 tampilkan
                print(f' {x} ; {cell.coordinate} ; {cell.value} ;')


print('\n===End Searching===;')
wbook.close()
